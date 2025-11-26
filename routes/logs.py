# routes/logs.py
from flask import Blueprint, render_template, request, abort, Response, send_file, jsonify, url_for
from io import BytesIO
import re

# Auth/session
from routes.auth import login_required, current_user
from services.auth_repository import user_can_view_tc

# DB helpers
from services.db import query_all, query_one
from services.tc_repository import list_tcs

# Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


logs_bp = Blueprint("logs", __name__, url_prefix="")

# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def _permitted_cts():
    u = current_user()
    if not u:
        abort(401)
    all_tcs = list_tcs()
    permitted = [c for c in all_tcs if user_can_view_tc(u, c["id"])]
    return u, permitted


def _get_session_or_404(session_id: int):
    s = query_one(
        """
        SELECT s.id, s.ct_id, s.lote, s.data_inicio, s.data_fim, s.status,
               s.total_final,
               s.contagem_alvo, s.observacao,
               c.name AS ct_name
          FROM session s
          JOIN tc c ON c.id = s.ct_id
         WHERE s.id = %s
        """,
        [session_id],
    )
    if not s:
        abort(404)
    _, permitted_cts = _permitted_cts()
    permitted_ids = {c["id"] for c in permitted_cts}
    if s["ct_id"] not in permitted_ids:
        abort(403)
    return s


def _get_session_logs(session_id: int):
    return query_all(
        """
        SELECT id, ts, delta, total_atual
          FROM session_log
         WHERE session_id = %s
         ORDER BY ts ASC
        """,
        [session_id],
    )


# Excel sheet titles cannot contain: : \ / ? * [ ]
# and must be <= 31 chars
def _safe_sheet_title(name: str) -> str:
    if not name:
        return "Planilha"
    bad = r'[:\\/*?\[\]/]'   # inclui a barra /
    cleaned = re.sub(bad, " ", str(name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return (cleaned or "Planilha")[:31]


def _safe_filename_piece(text: str) -> str:
    if not text:
        return ""
    piece = re.sub(r'[^0-9A-Za-z_\-]+', '_', text.strip())
    piece = re.sub(r'_+', '_', piece).strip('_')
    return piece


# ----------------------------------------------------------------------
# LISTA DE SESSÕES (com filtro por CT)
# ----------------------------------------------------------------------
@logs_bp.get("/sessoes")
@login_required
def logs_panel():
    """
    Mostra as sessões das CTs às quais o usuário tem acesso.
    Filtro opcional: ?ct_id=<id> ou ?ct_id=all
    """
    _, permitted_cts = _permitted_cts()
    if not permitted_cts:
        return render_template(
            "sessoes_panel.html",
            cts=[],
            current_ct_id="all",
            sessions=[],
            current_status="operando",
            current_lote="",
            current_ini_de="",
            current_ini_ate="",
            page=1,
            per=20,
            total_pages=1,
        )

    ct_id_arg = request.args.get("ct_id", "all")
    status_arg = request.args.get("status", "operando").strip().lower()
    lote_arg = (request.args.get("lote") or "").strip()
    ini_de = (request.args.get("ini_de") or "").strip()
    ini_ate = (request.args.get("ini_ate") or "").strip()
    # pagination controls (fixed 20 per page)
    per = 20
    try:
        page = int(request.args.get("page", "1"))
    except ValueError:
        page = 1
    if page < 1:
        page = 1
    permitted_ids = [c["id"] for c in permitted_cts]

    if ct_id_arg == "all":
        ct_ids = permitted_ids
        current_ct_id = "all"
    else:
        try:
            wanted = int(ct_id_arg)
        except ValueError:
            abort(400)
        if wanted not in permitted_ids:
            abort(403)
        ct_ids = [wanted]
        current_ct_id = str(wanted)

    placeholders = ",".join(["%s"] * len(ct_ids))
    where = [f"s.ct_id IN ({placeholders})"]
    params = list(ct_ids)
    if status_arg and status_arg != "all":
        where.append("s.status = %s")
        params.append(status_arg)
    if lote_arg:
        where.append("s.lote ILIKE %s")
        params.append(f"%{lote_arg}%")
    if ini_de:
        where.append("s.data_inicio::date >= %s")
        params.append(ini_de)
    if ini_ate:
        where.append("s.data_inicio::date <= %s")
        params.append(ini_ate)
    where_sql = " AND ".join(where)
    # total count for pagination
    sql_count = f"""
        SELECT COUNT(*) AS n
          FROM session s
          JOIN tc c ON c.id = s.ct_id
         WHERE {where_sql}
    """
    total_row = query_one(sql_count, params) or {"n": 0}
    total = int(total_row.get("n", 0))

    # page 1 => newest first, standard offset
    offset = (page - 1) * per

    sql = f"""
        SELECT
            s.id,
            s.ct_id,
            c.name     AS ct_name,
            s.lote,
            s.data_inicio,
            s.data_fim,
            s.status,
            s.contagem_alvo,
            COALESCE(s.total_final, 0) AS total_final,
            s.observacao
        FROM session s
          JOIN tc c ON c.id = s.ct_id
        WHERE {where_sql}
        ORDER BY s.data_inicio DESC
        OFFSET %s
        LIMIT %s
    """
    sessions = query_all(sql, params + [offset, per])

    total_pages = (total + per - 1) // per if per > 0 else 1
    if total_pages < 1:
        total_pages = 1

    # compute whether observation can be edited (finalizado and within 10 min)
    from datetime import datetime, timezone, timedelta
    now_local = datetime.now()
    now_utc = datetime.now(timezone.utc)
    for row in sessions:
        df = row.get("data_fim")
        st = (row.get("status") or "").lower()
        ok = False
        deadline_iso = None
        if st == "finalizado" and df is not None:
            try:
                if getattr(df, "tzinfo", None) is None:
                    deadline = df + timedelta(minutes=10)
                    diff = deadline - now_local
                    deadline_iso = deadline.isoformat()
                else:
                    deadline = df.astimezone(timezone.utc) + timedelta(minutes=10)
                    diff = deadline - now_utc
                    deadline_iso = deadline.isoformat()
                ok = diff.total_seconds() >= 0
            except Exception:
                ok = False
                deadline_iso = None
        row["can_edit_obs"] = ok
        row["obs_deadline_iso"] = deadline_iso
        row["obs_text"] = row.get("observacao") or ""

    return render_template(
        "sessoes_panel.html",
        cts=permitted_cts,
        current_ct_id=current_ct_id,
        sessions=sessions,
        current_status=status_arg,
        current_lote=lote_arg,
        current_ini_de=ini_de,
        current_ini_ate=ini_ate,
        page=page,
        per=per,
        total_pages=total_pages,
    )


# ----------------------------------------------------------------------
# DETALHE DA SESSÃO (logs)
# ----------------------------------------------------------------------
@logs_bp.get("/log/<int:session_id>")
@login_required
def log_detail(session_id: int):
    s = _get_session_or_404(session_id)
    
    # Exibição: por padrão mostra TODOS os eventos; opções: per=10, per=50
    per_arg = (request.args.get("per") or "all").strip().lower()
    show_all = per_arg == "all"
    per = 0 if show_all else (50 if per_arg == "50" else 10)
    try:
        page = int(request.args.get("page", "1"))
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    # Count total logs for session
    total_count_row = query_one(
        "SELECT COUNT(*) AS n FROM session_log WHERE session_id = %s",
        [session_id],
    ) or {"n": 0}
    total = int(total_count_row.get("n", 0))

    if show_all:
        rows = query_all(
            """
            SELECT id, ts, delta, total_atual
              FROM session_log
             WHERE session_id = %s
             ORDER BY id ASC
            """,
            [session_id],
        ) or []
        total_pages = 1
        page = 1
        per = total
    else:
        # Compute offset from the end so page 1 shows the latest 'per' rows
        offset = total - page * per
        if offset < 0:
            offset = 0
        rows = query_all(
            f"""
            SELECT id, ts, delta, total_atual
              FROM session_log
             WHERE session_id = %s
             ORDER BY id ASC
             OFFSET %s LIMIT %s
            """,
            [session_id, offset, per],
        ) or []
        total_pages = (total + per - 1) // per if per > 0 else 1
    if total_pages < 1:
        total_pages = 1

    # Effective total for header: prefer DB total_final if set, else last row
    effective_total = s.get("total_final") if s.get("total_final") is not None else (rows[-1]["total_atual"] if rows else 0)

    return render_template(
        "log_detail.html",
        sess=s,
        logs=rows,
        page=page,
        per=per,
        show_all=show_all,
        total=total,
        total_pages=total_pages,
        effective_total=effective_total,
    )

# Editar observação (somente até 10 minutos após finalizar)
@logs_bp.post("/log/<int:session_id>/observacao")
@login_required
def log_update_observacao(session_id: int):
    sess = _get_session_or_404(session_id)
    from datetime import datetime, timezone, timedelta
    df = sess.get("data_fim")
    st = (sess.get("status") or "").lower()
    if st != "finalizado" or not df:
        return ("Só é permitido alterar observação após finalizar.", 400)
    now_local = datetime.now()
    now_utc = datetime.now(timezone.utc)
    try:
        if getattr(df, "tzinfo", None) is None:
            deadline = df + timedelta(minutes=10)
            diff_seconds = (deadline - now_local).total_seconds()
        else:
            deadline = df.astimezone(timezone.utc) + timedelta(minutes=10)
            diff_seconds = (deadline - now_utc).total_seconds()
    except Exception:
        diff_seconds = -1
    if diff_seconds < 0:
        return ("Janela para alterar observação expirou.", 400)

    obs = (request.form.get("observacao") or "").strip()
    if len(obs) < 10:
        return ("Observação deve ter pelo menos 10 caracteres.", 400)
    from services.db import execute
    execute("UPDATE session SET observacao = %s WHERE id = %s", [obs, session_id])
    return ("", 204)

# Rota legada: redireciona para /log/<id>
@logs_bp.get("/logs/session/<int:session_id>")
@login_required
def legacy_session_logs(session_id: int):
    from flask import redirect, url_for
    return redirect(url_for('logs.log_detail', session_id=session_id), code=302)


# ----------------------------------------------------------------------
# LIVE JSON: novos eventos da sessão (para auto-atualização do detalhe)
# ----------------------------------------------------------------------
@logs_bp.get("/log/<int:session_id>/events.json")
@login_required
def log_events_json(session_id: int):
    sess = _get_session_or_404(session_id)
    try:
        after_id = int(request.args.get("after_id", 0))
    except Exception:
        after_id = 0

    rows = query_all(
        """
        SELECT id, ts, delta, total_atual
          FROM session_log
         WHERE session_id = %s AND id > %s
         ORDER BY id ASC
        """,
        [session_id, after_id],
    ) or []

    items = []
    for r in rows:
        ts_txt = r.get("ts").strftime("%H:%M:%S") if r.get("ts") else ""
        items.append({
            "id": r.get("id"),
            "ts": ts_txt,
            "delta": r.get("delta", 0),
            "total_atual": r.get("total_atual", 0),
        })

    return jsonify({"items": items})

# Rota legada: JSON events
@logs_bp.get("/logs/session/<int:session_id>/events.json")
@login_required
def legacy_session_logs_events_json(session_id: int):
    return log_events_json(session_id)


# ----------------------------------------------------------------------
# EXPORT XLSX — layout igual ao do seu exemplo
# ----------------------------------------------------------------------
@logs_bp.get("/log/<int:session_id>/export.xlsx")
@login_required
def log_export_xlsx(session_id: int):
    sess = _get_session_or_404(session_id)
    rows = _get_session_logs(session_id) or []

    total_final = sess.get("total_final", 0) or (rows[-1]["total_atual"] if rows else 0)

    wb = Workbook()
    ws = wb.active
    sheet_title = _safe_sheet_title(sess.get("ct_name") or f"TC {sess['ct_id']}")
    ws.title = sheet_title

    # Estilos
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="004A80")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho superior
    ws["A1"] = "TC";             ws["A1"].font = title_font
    ws["B1"] = sess.get("ct_name") or f"TC {sess['ct_id']}"
    ws["A2"] = "DATA INICIO";    ws["A2"].font = bold
    ws["B2"] = sess["data_inicio"].strftime("%d/%m/%Y %H:%M") if sess.get("data_inicio") else "-"
    ws["A3"] = "DATA FIM";       ws["A3"].font = bold
    ws["B3"] = sess["data_fim"].strftime("%d/%m/%Y %H:%M") if sess.get("data_fim") else "-"
    ws["A4"] = "TOTAL";          ws["A4"].font = bold
    ws["B4"] = total_final
    ws["A5"] = "CONTAGEM ALVO";  ws["A5"].font = bold
    ws["B5"] = sess.get("contagem_alvo") if sess.get("contagem_alvo") is not None else "-"
    ws["A6"] = "OBSERVACAO";     ws["A6"].font = bold
    ws["B6"] = sess.get("observacao") or "-"
    ws["B6"].alignment = wrap_left

    ws.append([])  # linha vazia

    # Tabela
    start_row = ws.max_row + 1
    headers = ["Status", "Hora", "Delta (+1/-1)", "Total Atual"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    r = start_row + 1
    for item in rows:
        status_txt = "RECONHECIMENTO"
        hora_txt = item["ts"].strftime("%H:%M:%S") if item.get("ts") else ""
        delta_val = item.get("delta", 0)
        total_val = item.get("total_atual", 0)

        ws.cell(row=r, column=1, value=status_txt).alignment = left
        ws.cell(row=r, column=2, value=hora_txt).alignment = center
        ws.cell(row=r, column=3, value=delta_val).alignment = center
        ws.cell(row=r, column=4, value=total_val).alignment = center

        for c in range(1, 5):
            ws.cell(row=r, column=c).border = border
        r += 1

    # larguras de coluna
    widths = {1: 18, 2: 14, 3: 16, 4: 14}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for rr in (1, 2, 3, 4, 5):
        ws[f"A{rr}"].alignment = left
        ws[f"B{rr}"].alignment = left
    ws["A6"].alignment = left
    ws.column_dimensions["A"].width = max(16, ws.column_dimensions["A"].width or 0)
    ws.column_dimensions["B"].width = max(24, ws.column_dimensions["B"].width or 0)

    # Nome do arquivo: session_ct<CTID>_<LOTE>.xlsx (sanitizado)
    ct_id = sess["ct_id"]
    lote = _safe_filename_piece(sess.get("lote") or "")
    filename = f"session_tc{ct_id}" + (f"_{lote}" if lote else "") + ".xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------------------------
# Export CSV simples (opcional)
# ----------------------------------------------------------------------
@logs_bp.get("/log/<int:session_id>/export.csv")
@login_required
def log_export_csv(session_id: int):
    sess = _get_session_or_404(session_id)
    rows = _get_session_logs(session_id)

    def gen_csv():
        yield "session_id,ct_id,ct_name,lote,ts,delta,total_atual\n"
        for r in rows:
            ts_str = r["ts"].strftime("%Y-%m-%d %H:%M:%S") if r.get("ts") else ""
            ct_name = (sess.get("ct_name") or "").replace(",", " ")
            lote = (sess.get("lote") or "").replace(",", " ")
            yield f"{sess['id']},{sess['ct_id']},{ct_name},{lote},{ts_str},{r.get('delta',0)},{r.get('total_atual',0)}\n"

    lote_piece = _safe_filename_piece(sess.get("lote") or "")
    filename = f"session_tc{sess['ct_id']}" + (f"_{lote_piece}" if lote_piece else "") + ".csv"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "text/csv; charset=utf-8",
    }
    return Response(gen_csv(), headers=headers)

# Rotas legadas: export CSV/XLSX
@logs_bp.get("/logs/session/<int:session_id>/export.xlsx")
@login_required
def legacy_session_logs_export_xlsx(session_id: int):
    return log_export_xlsx(session_id)

@logs_bp.get("/logs/session/<int:session_id>/export.csv")
@login_required
def legacy_session_logs_export_csv(session_id: int):
    return log_export_csv(session_id)
# Legacy index: /logs -> /sessoes (preserva querystring)
@logs_bp.get("/logs")
@login_required
def legacy_logs_index():
    from flask import redirect, request, url_for
    base = url_for('logs.logs_panel')
    qs = request.query_string.decode() if hasattr(request, 'query_string') else ''
    return redirect(base + (('?' + qs) if qs else ''), code=302)


# ----------------------------------------------------------------------
# EXPORT PDF — layout com múltiplas colunas (Hora | I | **)
# ----------------------------------------------------------------------
@logs_bp.get("/log/<int:session_id>/export.pdf")
@login_required
def log_export_pdf(session_id: int):
    sess = _get_session_or_404(session_id)
    rows = _get_session_logs(session_id) or []

    # Importa sob demanda e falha de modo amigável se não instalado
    try:
        from io import BytesIO
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
    except ModuleNotFoundError:
        return (
            "Dependência ausente para gerar PDF (reportlab). "
            "Instale no servidor central com: pip install -r requirements-central.txt",
            500,
            {"Content-Type": "text/plain; charset=utf-8"},
        )

    total_final = sess.get("total_final") if sess.get("total_final") is not None else (rows[-1]["total_atual"] if rows else 0)

    buf = BytesIO()
    # Página em retrato (A4)
    pw, ph = A4
    c = canvas.Canvas(buf, pagesize=(pw, ph))

    # Layout base
    margin_l, margin_r, margin_t, margin_b = 18, 18, 22, 26
    header_h1, header_h2, header_gap = 22, 18, 8
    footer_h = 80
    body_top = ph - margin_t - header_h1 - header_h2 - header_gap
    body_bottom = margin_b + footer_h
    usable_w = pw - margin_l - margin_r
    usable_h = body_top - body_bottom

    # Grupos de colunas em retrato: 11 grupos ("Hora | I | **")
    groups = 11
    gap = 4
    group_w = (usable_w - gap * (groups - 1)) / groups
    sub_w_hora = group_w * 0.55
    sub_w_delta = group_w * 0.17
    sub_w_total = group_w * 0.28

    # Alturas menores para caber mais linhas (retrato)
    row_h = 7
    header_row_h = 10
    rows_per_col_geom = int((usable_h - header_row_h) // row_h) or 1
    total_rows = len(rows)
    # 90 linhas POR COLUNA; calcula capacidade por página = 90 * ncol (limitada pela geometria)
    lines_per_col_target = 90
    rows_per_col_target = min(rows_per_col_geom, lines_per_col_target)
    page_capacity_total = rows_per_col_target * groups
    total_pages = max(1, (total_rows + page_capacity_total - 1) // page_capacity_total)

    def draw_header(page_no: int):
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#002F54"))
        c.drawString(margin_l, ph - margin_t, "Coonagro")
        title = f"RELATÓRIO DE CONTAGEM DE SACARIA" + (f" - {sess.get('ct_name')}" if sess.get('ct_name') else "")
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(pw/2, ph - margin_t, title)
        # Página em fonte menor
        c.setFont("Helvetica", 6)
        c.drawRightString(pw - margin_r, ph - margin_t, f"PÁGINA {page_no}/{total_pages}")
        # Linha abaixo do título
        c.setStrokeColor(colors.HexColor("#004A80"))
        c.setLineWidth(0.8)
        c.line(margin_l, ph - margin_t - 4, pw - margin_r, ph - margin_t - 4)

        lote = (sess.get("lote") or "-")
        qtde_alvo = (str(sess.get("contagem_alvo")) if sess.get("contagem_alvo") is not None else "-")
        inicio = sess["data_inicio"].strftime("%d/%m/%Y %H:%M:%S") if sess.get("data_inicio") else "-"
        fim = sess["data_fim"].strftime("%d/%m/%Y %H:%M:%S") if sess.get("data_fim") else "-"
        line2 = f"LOTE: {lote}  |  QTDE {qtde_alvo}  |  INÍCIO: {inicio} - FIM: {fim}  |  TOTAL CONTADO: {total_final}"
        c.setFont("Helvetica", 10)
        c.drawCentredString(pw/2, ph - margin_t - header_h1 - 2, line2)

    def draw_footer_if_last(page_no: int):
        if page_no != total_pages:
            return
        # Caixa de observação
        y0 = margin_b + 60
        c.setLineWidth(1)
        c.setStrokeColor(colors.black)
        c.rect(margin_l, y0, pw - margin_l - margin_r, 45, stroke=1, fill=0)
        obs = (sess.get("observacao") or "").strip()
        if obs:
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(pw/2, y0 + 15, obs[:3000])
        # Assinaturas
        c.setLineWidth(0.8)
        left_x = margin_l + 60
        right_x = pw - margin_r - 200
        base_y = margin_b + 20
        c.line(left_x, base_y, left_x + 240, base_y)
        c.line(right_x, base_y, right_x + 240, base_y)
        c.setFont("Helvetica", 9)
        c.drawCentredString(left_x + 120, base_y - 12, "NOME DO RESPONSÁVEL")
        c.drawCentredString(right_x + 120, base_y - 12, "ASSINATURA DO RESPONSÁVEL")

    def draw_table_page(start_index: int, page_no: int, remaining: int):
        draw_header_custom(page_no)
        # Cabeçalho dos grupos
        x = margin_l
        y_header = body_top
        c.setFillColor(colors.HexColor("#004A80"))
        c.setStrokeColor(colors.HexColor("#004A80"))
        for _g in range(groups):
            c.rect(x, y_header - header_row_h, group_w, header_row_h, fill=1, stroke=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 4, y_header - header_row_h + 3, "Hora")
            c.drawRightString(x + sub_w_hora + sub_w_delta - 4, y_header - header_row_h + 3, "I")
            c.drawRightString(x + group_w - 4, y_header - header_row_h + 3, "**")
            c.setFillColor(colors.HexColor("#004A80"))
            x += group_w + gap

        # Grade
        c.setStrokeColor(colors.HexColor("#D1D5DB"))
        c.setLineWidth(0.5)
        y0 = y_header - header_row_h
        # Linhas por coluna desta página: distribui o restante pelas colunas
        from math import ceil
        # Limita para distribuir no máximo 90 linhas por página
        eff_rows_per_col = min(
            rows_per_col_geom,
            ceil(min(remaining, page_capacity_target) / groups)
        ) or 1

        for r in range(eff_rows_per_col + 1):
            yy = y0 - r * row_h
            x = margin_l
            for _g in range(groups):
                c.line(x, yy, x + group_w, yy)
                x += group_w + gap
        x = margin_l
        for _g in range(groups):
            c.line(x, y0, x, y0 - eff_rows_per_col * row_h)
            c.line(x + group_w, y0, x + group_w, y0 - eff_rows_per_col * row_h)
            c.line(x + sub_w_hora, y0, x + sub_w_hora, y0 - eff_rows_per_col * row_h)
            c.line(x + sub_w_hora + sub_w_delta, y0, x + sub_w_hora + sub_w_delta, y0 - eff_rows_per_col * row_h)
            x += group_w + gap

        # Dados
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.black)
        for g in range(groups):
            for r in range(eff_rows_per_col):
                idx = start_index + g * eff_rows_per_col + r
                if idx >= total_rows:
                    continue
                item = rows[idx]
                ts_txt = item.get("ts").strftime("%H:%M:%S") if item.get("ts") else ""
                delta_txt = f"{item.get('delta', 0)}"
                total_txt = f"{item.get('total_atual', 0)}"
                base_x = margin_l + g * (group_w + gap)
                base_y = y0 - r * row_h
                c.drawString(base_x + 3, base_y - row_h + 3, ts_txt)
                c.drawRightString(base_x + sub_w_hora + sub_w_delta - 3, base_y - row_h + 3, delta_txt)
                c.drawRightString(base_x + group_w - 3, base_y - row_h + 3, total_txt)

        draw_footer_if_last(page_no)
        c.showPage()

    # Cabeçalho novo: título fixo, depois nome da TC, depois linha de LOTE
    def draw_header_custom(page_no: int):
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#002F54"))
        c.drawString(margin_l, ph - margin_t, "Coonagro")
        # Título fixo
        title = "RELATORIO DE CONTAGEM DE SACARIA"
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(pw/2, ph - margin_t, title)
        # Página (menor)
        c.setFont("Helvetica", 6)
        c.drawRightString(pw - margin_r, ph - margin_t, f"PAGINA {page_no}/{total_pages}")
        # Linha abaixo do título
        c.setStrokeColor(colors.HexColor("#004A80"))
        c.setLineWidth(0.8)
        c.line(margin_l, ph - margin_t - 4, pw - margin_r, ph - margin_t - 4)

        # Nome da TC em linha dedicada
        ct_txt = (sess.get("ct_name") or f"TC {sess['ct_id']}")
        c.setFont("Helvetica-Bold", 9)
        y_ct = ph - margin_t - header_h1 - 6
        c.drawCentredString(pw/2, y_ct, ct_txt)

        # Linha com LOTE/INICIO/FIM/TOTAL
        lote = (sess.get("lote") or "-")
        qtde_alvo = (str(sess.get("contagem_alvo")) if sess.get("contagem_alvo") is not None else "-")
        inicio = sess["data_inicio"].strftime("%d/%m/%Y %H:%M:%S") if sess.get("data_inicio") else "-"
        fim = sess["data_fim"].strftime("%d/%m/%Y %H:%M:%S") if sess.get("data_fim") else "-"
        line2 = f"LOTE: {lote}  |  QTDE {qtde_alvo}  |  INICIO: {inicio} - FIM: {fim}  |  TOTAL CONTADO: {total_final}"
        c.setFont("Helvetica", 9)
        y_line2 = y_ct - 10
        c.drawCentredString(pw/2, y_line2, line2)

    # Versão v2: distribui exatamente 90 itens por página entre 11 colunas,
    # com algumas colunas recebendo +1 linha quando necessário. Fonte menor nos dados.
    def draw_table_page_v2(start_index: int, page_no: int, remaining: int):
        draw_header(page_no)
        # Cabeçalho dos grupos
        x = margin_l
        y_header = body_top
        c.setFillColor(colors.HexColor("#004A80"))
        c.setStrokeColor(colors.HexColor("#004A80"))
        for _g in range(groups):
            c.rect(x, y_header - header_row_h, group_w, header_row_h, fill=1, stroke=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 2, y_header - header_row_h + 2, "Hora")
            c.drawRightString(x + sub_w_hora + sub_w_delta - 2, y_header - header_row_h + 2, "I")
            c.drawRightString(x + group_w - 2, y_header - header_row_h + 2, "**")
            c.setFillColor(colors.HexColor("#004A80"))
            x += group_w + gap

        # Grade com colunas de alturas variáveis (balanceado: round-robin entre colunas)
        c.setStrokeColor(colors.HexColor("#D1D5DB"))
        c.setLineWidth(0.5)
        y0 = y_header - header_row_h
        # Quantos itens vamos renderizar nesta página ao todo
        target_this_page = min(remaining, page_capacity_total)
        # Distribuição round-robin: base + 1 nas primeiras 'rem' colunas
        base = target_this_page // groups
        rem = target_this_page % groups
        col_rows = [min(rows_per_col_target, base + (1 if g < rem else 0)) for g in range(groups)]

        x = margin_l
        for g in range(groups):
            rows_g = col_rows[g]
            for r in range(rows_g + 1):
                yy = y0 - r * row_h
                c.line(x, yy, x + group_w, yy)
            c.line(x, y0, x, y0 - rows_g * row_h)
            c.line(x + group_w, y0, x + group_w, y0 - rows_g * row_h)
            c.line(x + sub_w_hora, y0, x + sub_w_hora, y0 - rows_g * row_h)
            c.line(x + sub_w_hora + sub_w_delta, y0, x + sub_w_hora + sub_w_delta, y0 - rows_g * row_h)
            x += group_w + gap

        # Dados com fonte menor
        c.setFont("Helvetica", 5)
        c.setFillColor(colors.black)
        offsets = []
        _acc = 0
        for g in range(groups):
            offsets.append(_acc)
            _acc += col_rows[g]
        for g in range(groups):
            rows_g = col_rows[g]
            for r in range(rows_g):
                idx = start_index + offsets[g] + r
                if idx >= total_rows:
                    continue
                item = rows[idx]
                ts_txt = item.get("ts").strftime("%H:%M:%S") if item.get("ts") else ""
                delta_txt = f"{item.get('delta', 0)}"
                total_txt = f"{item.get('total_atual', 0)}"
                base_x = margin_l + g * (group_w + gap)
                base_y = y0 - r * row_h
                c.drawString(base_x + 2, base_y - row_h + 2, ts_txt)
                c.drawRightString(base_x + sub_w_hora + sub_w_delta - 2, base_y - row_h + 2, delta_txt)
                c.drawRightString(base_x + group_w - 2, base_y - row_h + 2, total_txt)

        draw_footer_if_last(page_no)
        c.showPage()

    if total_rows == 0:
        draw_header_custom(1)
        draw_footer_if_last(1)
        c.showPage()
    else:
        start = 0
        page_no = 1
        while start < total_rows:
            remaining = total_rows - start
            draw_table_page_v2(start, page_no, remaining)
            # Avança respeitando o limite de 90 linhas por página
            step = min(remaining, page_capacity_total)
            start += step
            page_no += 1

    c.save()
    buf.seek(0)

    ct_id = sess["ct_id"]
    lote = _safe_filename_piece(sess.get("lote") or "")
    filename = f"session_tc{ct_id}" + (f"_{lote}" if lote else "") + ".pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")
